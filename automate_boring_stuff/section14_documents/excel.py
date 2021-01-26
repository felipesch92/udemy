import openpyxl

workbook = openpyxl.load_workbook('Controle Familiar 2021.xlsx')
sheet = workbook['Janeiro']
print(sheet)
print(workbook.sheetnames)
ell = sheet['B2']
print(sheet['B2'].value)

for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)
